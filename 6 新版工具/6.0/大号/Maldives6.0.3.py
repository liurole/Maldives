# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 16:44:45 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts

记得改成自己的账户密码！！！

实现以下功能：账号登录，Excel读取，系列产品汇总，requests查重，选型主题提交
"""
import time
import openpyxl
from selenium import webdriver

# 此函数实现了Excel数据表的提前，默认第一列为表头，同时可以自己选择第几个表
def excel_table_byindex(file, index = 0):  
    # 找到目标sheet
    wb = openpyxl.reader.excel.load_workbook(file)
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[index])

    # 获取表头，这里第一行为表头，如果不是，此程序不能用
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数    
    results = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        results.append(app)  
    print('Read Done!')    
 
    return results 

# 此函数实现了相同序号的汇总，主要针对于系列产品分类
def sort_table(results):
    sorted_results = {}
    index = 1
    temp = []
    ids = []
    for i in results:
        ids.append(i['序号'])
    nums = len(set(ids))
    
    while index != nums + 1:
        for i, val in enumerate(results):
            if index == val['序号']:   
                temp.append(val)
            if i == len(results) - 1:
                sorted_results[index] = temp
                index += 1   
                temp = []
    return sorted_results

# 提交单个系列产品，仅适用于新产品
def submit(browser, data, submit_url):
    urls = []
    versions = ''
    # 主要是获取系列型号的信息汇总
    init = True
    for val  in data:
        urls.append(val['参考链接'])
        if init:
            versions = val['型号']
            init = False
        else:
            versions += ',' + val['型号']
    print('涉及型号\t' + versions)
    
    browser.get(submit_url)
    time.sleep(2)
    # 参考链接一个是翻译，多个是改写(大幅度的整理以及自己的分析整理)
    browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[1]/div/div/div/div[2]/div/span/div/label[3]/span[1]').click()
    time.sleep(2)
    
    # 原文链接
    url = browser.find_element_by_id('refUrls_0')
    url.send_keys(urls[0])
    url = browser.find_element_by_id('refUrls_1')
    url.send_keys(urls[1])
    for i, val in enumerate(urls):
        if i > 1:
            temp = '//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[2]/div[' + str(i + 1) + ']/div/div/div/div/div/span/button'
            browser.find_element_by_xpath(temp).click()
            temp = 'refUrls_' + str(i)
            url = browser.find_element_by_id(temp)
            url.send_keys(val)
    
    # 厂牌（这里选京瓷）
    if data[0]['厂牌'] == 'Kyocera(京瓷)':
        browser.find_element_by_id('brandCode').click()
        time.sleep(2)
        browser.find_element_by_id('brandCodetreeLeft_47_span').click()
        browser.find_element_by_id('brandCodeconfirmBtn').click()   
    elif data[0]['厂牌'] == 'shindengen（新电元）':
        browser.find_element_by_id('brandCode').click()
        time.sleep(2)
        browser.find_element_by_id('brandCodetreeLeft_24_span').click()
        browser.find_element_by_id('brandCodeconfirmBtn').click()           
    
    # 器件名称
    name = browser.find_element_by_id('goodKeyword')
    name.send_keys('液晶显示屏,TFT,薄膜晶体管')
    
    # 型号
    name = browser.find_element_by_id('pnKeyword')
    name.send_keys(versions)
    
    # 理由
    reason_out = '未有雷同文章，所涉及型号均为2010年以后产品，由自己总结改写，主题为：' + data[0]['选型应用']
    reason = browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[6]/div/div/div/div[2]/div/span/textarea')
    reason.send_keys(reason_out)
    
    # 选型类型
    browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[7]/div/div/div/div[2]/div/span/div/label[5]/span[1]/input').click()
    
    # 查重
    browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[8]/div/div/div/div/div/span/button[1]').click()
    
    pick = input("是否已手动提交？ y/n:")
    if pick == 'y':
#        # 提交
#        browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/div/form/div[8]/div/div/div/div/div/button[2]').click()
        print('next')
        return 1
    else:
        return 404
        
    
if __name__ == '__main__':
    
    # 打开chorme
    url = 'https://cms.sekorm.com/content/login'
    browser = webdriver.Chrome()
    browser.get(url)
    time.sleep(1)
    
    # 输入验证码
    name = browser.find_element_by_id('mobile')
    name.send_keys('15244608508')
    password = browser.find_element_by_id('password')
    password.send_keys('liu875288')
    button = browser.find_element_by_xpath('//*[@id="root"]/div/div/div[2]/div/div[1]/div/div/div/div[2]/form/div[3]/div/button')
    button.click()
    
    # 切换到提主题页面
    url = 'https://cms.sekorm.com/content/nps/selfSubject/add'
    
    # 读取Excel，同时合并同一序号，序号编号需要连续
    file = '选型.xlsx'
    results = excel_table_byindex(file, 3)
    sorted_results = sort_table(results)

    # 依次提交
    for key, value in sorted_results.items():
        if key > 14:
            state = submit(browser, value, url)
            if state == 404:
                break
               
