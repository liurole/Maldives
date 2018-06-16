# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 16:44:45 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts

记得改成自己的账户密码！！！

实现以下功能：账号登录，Excel读取，系列产品汇总，requests查重，新产品系列产品提交
"""
import time
import openpyxl
from selenium import webdriver
import requests
import re
from bs4 import BeautifulSoup
from urllib.parse import urlencode

# 此函数实现了Excel数据表的提前，默认第一列为表头，同时可以自己选择第几个表
def excel_table_byindex(file, index = 0):  
    # 找到目标sheet
    wb = openpyxl.reader.excel.load_workbook(file)
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[index])

    # 获取表头，这里第一行为表头，如果不是，此程序不能用
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数    
    results = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        results.append(app)  
    print('Read Done!')    
 
    return results 

# 此函数实现了相同序号的汇总，主要针对于系列产品分类
def sort_table(results):
    sorted_results = {}
    index = 1
    temp = []
    ids = []
    for i in results:
        ids.append(i['序号'])
    nums = len(set(ids))
    
    while index != nums + 1:
        for i, val in enumerate(results):
            if index == val['序号']:   
                temp.append(val)
            if i == len(results) - 1:
                sorted_results[index] = temp
                index += 1   
                temp = []
    return sorted_results

# 利用request，导入cookies，header进行关键词网页搜索，选择第二栏
def check_repeated(keyword, tab = 1):
    headers = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate, sdch, br',
    'Accept-Language': 'zh-CN,zh;q=0.8',
    'Connection': 'keep-alive',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.79 Safari/537.36 Edge/14.14393',
    'X-Requested-With': 'XMLHttpRequest'
    }
    data = {
        'tab': str(tab),
    }
    params = urlencode(data)
    base = 'https://www.sekorm.com/Web/Search/keyword/'
    url = base + keyword + '?' + params
    try:
        response = requests.get(url, headers = headers)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

# 获取所有搜索结果链接
def get_all_urls(html):
    names = []
    urls = []
    soup = BeautifulSoup(html, 'lxml')
    result = soup.select('#multiple-search_0 a')
    if len(result) == 0:
        return names, urls
    else:
        for i in result:
            temp = i.get('href')
            urls.append(temp)
            temp = i.get_text()
            names.append(temp)
        return names, urls

# 搜索所有子链接，在子页面进行查找，发现是否包括此关键词
def search_all_urls(urls, keyword, length = 20):
    length = min(length, len(urls))
    for i in range(length):
        html = urls[i]
        response = requests.get(html)
        soup = BeautifulSoup(response.text, 'lxml')
        result = soup.select('.cd-content p')
        for j in result:
            temp = result = re.sub(r'[^\x00-\x7F]+', '', j.get_text())
            if keyword in temp:
                return True
    return False

# 查重汇总程序
def is_repeated(keyword):
    html = check_repeated(keyword)
    names, urls = get_all_urls(html)
    repeated = search_all_urls(urls, keyword)  
    return repeated

# 提交单个系列产品，仅适用于新产品
def submit(browser, data, url):
    browser.get(url)
    urls = ''
    versions = ''
    dates = ''
    # 主要是获取系列型号的信息汇总
    init = True
    for val  in data:
        repeated = is_repeated(val['型号'])
        if not repeated:
            if init:
                urls = val['参考链接']
                versions = val['型号']
                dates = str(val['判断为新文章的理由'])
                init = False
            else:
                urls += ',' + val['参考链接']
                versions += ',' + val['型号']
                dates += '/' + str(val['判断为新文章的理由'])
        else:
            print(val['型号'] + '\trepeated' + '\t请留意')
    if versions == '':
        return -1
    dates = '数据手册形成时间为' + str(dates) + '，世强官网尚未有此型号文章。'
    print('型号\t' + versions)
    print('形成时间\t' + dates)
    
    # 参考链接一个是翻译，多个是改写(大幅度的整理以及自己的分析整理)
    browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[1]/div/div/div/div[2]/div/span/div/label[2]/span[1]/input').click()
    
    # 原文链接
    url = browser.find_element_by_id('originalUrl')
    url.send_keys(urls)
    
    # 厂牌（这里选京瓷）
    if data[0]['厂牌'] == 'Kyocera(京瓷)':
        browser.find_element_by_id('brandCode').click()
        time.sleep(2)
        browser.find_element_by_id('brandCodetreeLeft_47_span').click()
        browser.find_element_by_id('brandCodeconfirmBtn').click()   
    elif data[0]['厂牌'] == 'shindengen（新电元）':
        browser.find_element_by_id('brandCode').click()
        time.sleep(2)
        browser.find_element_by_id('brandCodetreeLeft_24_span').click()
        browser.find_element_by_id('brandCodeconfirmBtn').click()           
    
    # 器件名称
    name = browser.find_element_by_id('goodKeyword')
    name.send_keys(data[0]['器件名称'])
    
    # 型号
    name = browser.find_element_by_id('pnKeyword')
    name.send_keys(versions)
    
    # 理由
    reason = browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[6]/div/div/div/div[2]/div/span/textarea')
    reason.send_keys(dates)
    
    # 新产品类型
    browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[7]/div/div/div/div[2]/div/span/div/label[1]/span[1]/input').click()
    
    # 查重
    browser.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[2]/div[2]/div/div/div/div/form/div[8]/div/div/div/div/div/span/button[1]').click()
        
    return 1
    
if __name__ == '__main__':
    
    # 打开chorme
    url = 'https://cms.sekorm.com/content/login'
    browser = webdriver.Chrome()
    browser.get(url)
    
    # 输入验证码
    name = browser.find_element_by_id('mobile')
    name.send_keys('15244608508')
    password = browser.find_element_by_id('password')
    password.send_keys('liu875288')
    button = browser.find_element_by_xpath('//*[@id="root"]/div/div/div[2]/div/div[1]/div/div/div/div[2]/form/div[3]/div/button')
    button.click()
    
    # 切换到提主题页面
    url = 'https://cms.sekorm.com/content/nps/selfSubject/add'
    
    # 读取Excel，同时合并同一序号，序号编号需要连续
    file = '提交.xlsx'
    results = excel_table_byindex(file, 4)
    sorted_results = sort_table(results)

    # 依次提交
    for value in sorted_results.values():
        state = submit(browser, value, url)
        pick = input("是否已手动提交？ y/n:")
        if pick == 'y':
            print('next')
        else:
            break

               
