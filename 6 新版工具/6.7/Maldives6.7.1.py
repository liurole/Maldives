# -*- coding: utf-8 -*-
"""
Created on Sat Dec  9 21:28:47 2017

@author: Se

马尔代夫V6.3.1

实现了对于指定网页集合的批量信息抓取

"""
import random
import openpyxl

# 此函数实现了Excel数据表的提前，默认第一列为表头，同时可以自己选择第几个表
def excel_table_byindex(file, index = 0):  
    # 找到目标sheet
    wb = openpyxl.reader.excel.load_workbook(file)
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[index])

    # 获取表头，这里第一行为表头，如果不是，此程序不能用
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数    
    results = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        results.append(app)  
    print('Read Done!')    
 
    return results 

if __name__ == '__main__':
    
    file = '每次提交后状态.xlsx'
    results = excel_table_byindex(file, 0)
    num = 0
    data = []
    ids = []
    while num < 30:
        temp_num = random.randint(0, len(results) - 1)
        ref_data = []
        is_empty = 0
        for i in results:
            if i['post'] != 1:
                if i['类型'] == temp_num:
                    if i['url'] == '':
                        is_empty = 1
                    ref_data.append(i)
                    ids.append(i['序号'])
                    i['post'] = 1
        if len(ref_data) != 0 and is_empty == 0:
            data.append(ref_data)
            num += 1
    
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '30篇提交'
    ws_out['A1'] = '序号'
    ws_out['B1'] = '类型'
    ws_out['C1'] = 'name'
    ws_out['D1'] = 'VRRM'
    ws_out['E1'] = 'IO'
    ws_out['F1'] = 'IFSM'
    ws_out['G1'] = 'IR'
    ws_out['H1'] = 'VFM'
    ws_out['I1'] = 'Tstg'
    ws_out['J1'] = 'Outline'
    ws_out['K1'] = 'Circuit'
    ws_out['L1'] = 'AEC-Q101'
    ws_out['M1'] = 'State'
    ws_out['N1'] = 'url'
    ws_out['O1'] = '原序号'  
    
    index = 2
    types = 1
    for j in data:
        for i in j:
            temp = 'A' + str(index)
            ws_out[temp] = index - 1
            temp = 'B' + str(index)
            ws_out[temp] = types
            temp = 'C' + str(index)
            ws_out[temp] = i['name']
            temp = 'D' + str(index)
            ws_out[temp] = i['VRRM']
            temp = 'E' + str(index)
            ws_out[temp] = i['IO']
            temp = 'F' + str(index)
            ws_out[temp] = i['IFSM']
            temp = 'G' + str(index)
            ws_out[temp] = i['IR']
            temp = 'H' + str(index)
            ws_out[temp] = i['VFM']
            temp = 'I' + str(index)
            ws_out[temp] = i['Tstg']
            temp = 'J' + str(index)
            ws_out[temp] = i['Outline']
            temp = 'K' + str(index)
            ws_out[temp] = i['Circuit']
            temp = 'L' + str(index)
            ws_out[temp] = i['AEC-Q101']
            temp = 'M' + str(index)
            ws_out[temp] = i['State']
            temp = 'N' + str(index)
            ws_out[temp] = i['url']
            temp = 'O' + str(index)
            ws_out[temp] = ids[index - 2]
            index += 1
        types += 1
        
    wb_out.save('30篇提交.xlsx')  
    

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '京瓷肖特基'
    ws_out['A1'] = '序号'
    ws_out['B1'] = '类型'
    ws_out['C1'] = 'name'
    ws_out['D1'] = 'VRRM'
    ws_out['E1'] = 'IO'
    ws_out['F1'] = 'IFSM'
    ws_out['G1'] = 'IR'
    ws_out['H1'] = 'VFM'
    ws_out['I1'] = 'Tstg'
    ws_out['J1'] = 'Outline'
    ws_out['K1'] = 'Circuit'
    ws_out['L1'] = 'AEC-Q101'
    ws_out['M1'] = 'State'
    ws_out['N1'] = 'url'
    ws_out['O1'] = 'post'
    
    index = 2
    for i in results:
        temp = 'A' + str(index)
        ws_out[temp] = i['序号']
        temp = 'B' + str(index)
        ws_out[temp] = i['类型']
        temp = 'C' + str(index)
        ws_out[temp] = i['name']
        temp = 'D' + str(index)
        ws_out[temp] = i['VRRM']
        temp = 'E' + str(index)
        ws_out[temp] = i['IO']
        temp = 'F' + str(index)
        ws_out[temp] = i['IFSM']
        temp = 'G' + str(index)
        ws_out[temp] = i['IR']
        temp = 'H' + str(index)
        ws_out[temp] = i['VFM']
        temp = 'I' + str(index)
        ws_out[temp] = i['Tstg']
        temp = 'J' + str(index)
        ws_out[temp] = i['Outline']
        temp = 'K' + str(index)
        ws_out[temp] = i['Circuit']
        temp = 'L' + str(index)
        ws_out[temp] = i['AEC-Q101']
        temp = 'M' + str(index)
        ws_out[temp] = i['State']
        temp = 'N' + str(index)
        ws_out[temp] = i['url']
        temp = 'O' + str(index)
        ws_out[temp] = i['post']
        
        index += 1

    wb_out.save('每次提交后状态.xlsx')   
    