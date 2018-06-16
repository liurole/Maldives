# -*- coding: utf-8 -*-
"""
Created on Sat Dec  9 21:28:47 2017

@author: Se


实现了批量pdf下载

"""

import openpyxl
import urllib

# 此函数实现了Excel数据表的提前，默认第一列为表头，同时可以自己选择第几个表
def excel_table_byindex(file, index = 0):  
    # 找到目标sheet
    wb = openpyxl.reader.excel.load_workbook(file)
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[index])

    # 获取表头，这里第一行为表头，如果不是，此程序不能用
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数    
    results = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        results.append(app)  
    print('Read Done!')    
 
    return results 


def getFile(url):
    file_name = url.split('/')[-1]
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}  
    req = urllib.request.Request(url, headers=headers) 
    data = urllib.request.urlopen(req).read()  
    f = open(file_dir + file_name, 'wb') 
    f.write(data)  
    f.close()
    print ("Sucessful to download" + " " + file_name)


# 帮助run Maldives3.0.py --help
if __name__ == '__main__':
    
    file = '京瓷肖特基提交整理版.xlsx'
    file_dir = './pdf/'
    results = excel_table_byindex(file, 0)
    
    for i in results:
        temp1 = i['name']
        temp2 = i['url']
        if temp2 != '':
            try:
                getFile(temp2)
            except:
                pass
            
    print('All done!')