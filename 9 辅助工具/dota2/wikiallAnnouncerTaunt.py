# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 16:44:45 2018

@author: Se

"""
import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook

def get_info(url):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

def get_data1(html): 
    names = []
    soup = BeautifulSoup(html, 'lxml')
    result = soup.select('#mw-content-text img') 
    index = 1
    for i in result:
        temp = str(i)
        name =re.search('img alt="(.*?)" height="', temp, re.S)
        if name == None:
            print(index)
        else:
            names.append(name.group(1))
        index += 1
        
    return names
    
def get_data2(html): 
    names = []
    soup = BeautifulSoup(html, 'lxml')
    result = soup.select('.cosmetic-label img') 
    index = 1
    for i in result:
        temp = str(i)
        name =re.search('img alt="(.*?)" height="', temp, re.S)
        if name == None:
            print(index)
        else:
            names.append(name.group(1))
        index += 1
        
    return names

if __name__ == '__main__':
    
    url = 'https://dota2.gamepedia.com/Announcers'
    html = get_info(url)
    names1 = get_data1(html)
    
    url = 'https://dota2.gamepedia.com/Taunt_(Equipment)'
    html = get_info(url)
    names2 = get_data2(html)
    
    
    index = 1
    wb = Workbook()
    ws = wb.active
    ws.title = '刀塔所有HUD名称'
    ws['A1'] = '序号'
    ws['B1'] = '类型'
    ws['C1'] = '物品名称'
    
    for i in names1:
        temp = 'A' + str(index + 1)
        ws[temp] = index
        temp = 'B' + str(index + 1)
        ws[temp] = 0  
        temp = 'C' + str(index + 1)
        ws[temp] = i 
        index += 1
    
    for i in names2:
        temp = 'A' + str(index + 1)
        ws[temp] = index
        temp = 'B' + str(index + 1)
        ws[temp] = 1  
        temp = 'C' + str(index + 1)
        ws[temp] = i 
        index += 1
        
    wb.save('刀塔所有AnnouncerTaunt名称.xlsx')
    
