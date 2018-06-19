# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 16:44:45 2018

@author: Se

"""
import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook

def get_info():
    url = 'https://dota2.gamepedia.com/Custom_Courier'
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

def get_data(html): 
    names = []
    soup = BeautifulSoup(html, 'lxml')
    result = soup.select('#mw-content-text img') 
    index = 1
    for i in result:
        temp = str(i)
        name =re.search('img alt="(.*?)" height="6', temp, re.S)
        if name == None:
            print(index)
        else:
            names.append(name.group(1))
        index += 1
        
    return names
    
    
if __name__ == '__main__':
    
    # 入口页面
    html = get_info()
    # 获取所有名称
    names = get_data(html)
    
    index = 1
    wb = Workbook()
    ws = wb.active
    ws.title = '刀塔所有信使名称'
    ws['A1'] = '序号'
    ws['B1'] = '物品名称'
    
    for i in names:
        temp = 'A' + str(index + 1)
        ws[temp] = index
        temp = 'B' + str(index + 1)
        ws[temp] = i        
        index += 1
            
    wb.save('刀塔所有信使名称.xlsx')
    
