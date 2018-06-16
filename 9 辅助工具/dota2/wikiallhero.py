# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 16:44:45 2018

@author: Se

"""
import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook

def get_info():
    url = 'https://dota2.gamepedia.com/Equipment'
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

def cut(html):
    result =re.search('<div class="heroentry">(.*?)</td></tr></table>', html, re.S)
    content = result.group(1)
    return content

def get_links(content):
    link = re.findall('<a href="(.*?)" title', content, re.S)
    links = []
    names = []
    for i in link:
        if 'Equipment' in i:
            links.append('https://dota2.gamepedia.com' + i)
            temp = i.split('/')
            names.append(temp[1])
    return links, names
    
def read_link(url):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None
    
#def get_data(html):    
#    sets = []
#    data = []
#    result =re.search('Sets</td>(.*?)</div></div></td></tr>', html, re.S)
#    if result == None:
#        return -1
#    else:
#        setsdata = result.group(1)
#        temp = re.findall('<a href="(.*?)" title', setsdata, re.S)
#        index = 0
#        for i in temp:
#            if index % 2 == 0:
#                sets.append(re.sub('_', ' ', i))
#            index += 1
#    
#    soup = BeautifulSoup(html, 'lxml')
#    result = soup.select('.cosmetic-label a')  
#    index = 0
#    for i in result:
#        temp = i.get('title')
#        if index % 2 == 0:
#            data.append(temp)
#        index += 1
    
def get_data(html):    
    sets = []
    data = []
    soup = BeautifulSoup(html, 'lxml')
    result = soup.select('.cosmetic-label a')  
    index = 0
    for i in result:
        temp = i.get('title')
        if index % 2 == 0:
            data.append(temp)
        index += 1    
    
    result =re.search('Sets</td>(.*?)</div></div></td></tr>', html, re.S)
    if result != None:
        setsdata = result.group(1)
        soup = BeautifulSoup(setsdata, 'lxml')
        result = soup.select('.cosmetic-label a')  
        index = 0
        for i in result:
            temp = i.get('title')
            if index % 2 == 0:
                sets.append(temp)
            index += 1  
    return sets, data
    
def collect_data(sets, data):
    isSet = []
    for i in data:
        if i in sets:
            isSet.append(1)
        else:
            isSet.append(0)
    return isSet
    
    
if __name__ == '__main__':
    
    # 入口页面
    html = get_info()
    # 提取与表格相关的内容
    content = cut(html)
    # 获取所有的英雄名称以及对应的链接
    links, names = get_links(content)
    
    total = []
    for index, url in enumerate(links):
        temp = {}
        html = read_link(url)
        sets, data = get_data(html)
        isSet = collect_data(sets, data)
        if len(data) == 0:
            print('!!!!NO DATA!!!!')
        temp['hero'] = names[index]
        temp['data'] = data
        temp['isSet'] = isSet
        total.append(temp)
        print(str(index) + '----' + names[index] + '----Done!')
    
    index = 1
    wb = Workbook()
    ws = wb.active
    ws.title = '刀塔所有英雄饰品名称'
    ws['A1'] = '序号'
    ws['B1'] = '子序号'
    ws['C1'] = '英雄'
    ws['D1'] = '物品名称'
    ws['E1'] = '是否为套装'
    
    for i in total:
        for ids, name in enumerate(i['data']):
            temp = 'A' + str(index + 1)
            ws[temp] = index
            temp = 'B' + str(index + 1)
            ws[temp] = ids + 1
            temp = 'C' + str(index + 1)
            ws[temp] = i['hero']
            temp = 'D' + str(index + 1)
            ws[temp] = name
            temp = 'E' + str(index + 1)
            ws[temp] = i['isSet'][ids]
            index += 1
            
    wb.save('刀塔所有英雄饰品名称.xlsx')
    
