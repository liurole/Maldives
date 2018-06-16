# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 16:44:45 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts

实现了利用本地cookie文件的登录
"""

import requests
import re
import math
from openpyxl import Workbook

def get_data(assetid, isInit = True, step = 1000):
    headers = {
    'Cache-Control': 'no-cache',
    'Connection': 'Transfer-Encoding',
    'Connection': 'keep-alive',
    'Content-Security-Policy': 'default-src blob: data: https: \'unsafe-inline\' \'unsafe-eval\'; script-src \'self\' \'unsafe-inline\' \'unsafe-eval\' https://steamcommunity-a.akamaihd.net/ https://api.steampowered.com/ *.google-analytics.com https://www.google.com https://www.gstatic.com https://apis.google.com; object-src \'none\'; connect-src \'self\' https://api.steampowered.com/ https://store.steampowered.com/ wss://community.steam-api.com/websocket/ *.google-analytics.com http://127.0.0.1:27060 ws://127.0.0.1:27060; frame-src \'self\' steam: https://store.steampowered.com/ https://www.youtube.com https://www.google.com https://sketchfab.com https://player.vimeo.com;',
    'Content-Type': 'application/json; charset=utf-8',
    'Date': 'Sat, 26 May 2018 13:09:57 GMT',
    'Expires': 'Mon, 26 Jul 1997 05:00:00 GMT',
    'Server': 'Apache',
    'Strict-Transport-Security': 'max-age=3600',
    'Transfer-Encoding': 'chunked',
    'X-Frame-Options': 'SAMEORIGIN'
    }
    pre_url = 'https://steamcommunity.com/inventory/76561198129439885/570/2?l=schinese&count='
    if isInit:
        url = pre_url + str(step)
    else:
        url = pre_url + str(step) + '&start_assetid=' + assetid
    try:
        response = requests.get(url, headers = headers)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None
        
def cut(html):
    result =re.search('descriptions":(.*)', html, re.S)
    descriptions = result.group()
    result =re.search('last_assetid":"(.*?)","total_inventory_count', html, re.S)
    if result == None:
        last_assetid = '-1'
    else:
        last_assetid = result.group(1)
    
    return descriptions, last_assetid

def get_descriptions(descriptions):
    currency = re.findall('currency":(.*?),"background_color', descriptions, re.S)
    tradable = re.findall('tradable":(.*?),"name', descriptions, re.S)
    name = re.findall('"name":"(.*?)","name_color', descriptions, re.S)
    market_name = re.findall('market_name":"(.*?)","market_hash_name', descriptions, re.S)
    market_hash_name = re.findall('market_hash_name":"(.*?)","commodity', descriptions, re.S)
    commodity = re.findall('commodity":(.*?),"market_tradable_restriction', descriptions, re.S)
    market_tradable_restriction = re.findall('market_tradable_restriction":(.*?),"market_marketable_restriction', descriptions, re.S)
    market_marketable_restriction = re.findall('market_marketable_restriction":(.*?),"marketable', descriptions, re.S)
    marketable = re.findall('marketable":(.*?),"tags', descriptions, re.S)
    
    quality = re.findall('质","localized_tag_name":"(.*?)","color', descriptions, re.S)
    rarity = re.findall('稀有度","localized_tag_name":"(.*?)","color', descriptions, re.S)
    types = re.findall('类型","localized_tag_name":"(.*?)"}', descriptions, re.S)
    slot = re.findall('槽位","localized_tag_name":"(.*?)"}', descriptions, re.S)
    hero = re.findall('英雄","localized_tag_name":"(.*?)"}', descriptions, re.S)
    
    temp = {}
    temp['currency'] = currency
    temp['tradable'] = tradable
    temp['name'] = name
    temp['market_name'] = market_name
    temp['market_hash_name'] = market_hash_name
    temp['commodity'] = commodity
    temp['market_tradable_restriction'] = market_tradable_restriction
    temp['market_marketable_restriction'] = market_marketable_restriction
    temp['marketable'] = marketable
    
    temp['quality'] = quality
    temp['rarity'] = rarity
    temp['types'] = types
    temp['slot'] = slot
    temp['hero'] = hero
    
    return temp


if __name__ == '__main__':
    
    step = 1000
    isInit = True
    assetid = 'start'
    num = 0
    result = []
    
    while True:
        if isInit:
            html = get_data(assetid, isInit, step)
            if html == None:
                print('运行错误！')
                break
            else:
                descriptions, last_assetid = cut(html)
                descriptions = get_descriptions(descriptions)
                str_out = str(num * step) + ' to ' +  str(num * step + step) + 'Done!'
                print(str_out)
                result.append(descriptions)
                assetid = last_assetid
                isInit = False
            num += 1
        else:
            html = get_data(assetid, isInit, step)
            if html == None:
                print('运行错误！')
                break
            else:
                descriptions, last_assetid = cut(html)
                descriptions = get_descriptions(descriptions)
                str_out = str(num * step) + ' to ' +  str(num * step + step) + 'Done!'
                print(str_out)
                result.append(descriptions)
                assetid = last_assetid
                if assetid == '-1':
                    break   
                num += 1
    
    index = 1
    wb = Workbook()
    ws = wb.active
    ws.title = '刀塔在线库存'
    ws['A1'] = '序号'
    ws['B1'] = '页数'
    ws['C1'] = '子序号'
    ws['D1'] = '名称'
    ws['E1'] = '市场中文名称'
    ws['F1'] = '市场英文名称'
    ws['G1'] = '品质'
    ws['H1'] = '稀有度'
    ws['I1'] = '英雄'
    ws['J1'] = '类型'
    ws['K1'] = '槽位'
    ws['L1'] = '可交易'
    ws['M1'] = '可出售'
    ws['N1'] = 'currency'
    ws['O1'] = 'commodity'
    ws['P1'] = 'market_tradable_restriction'
    ws['Q1'] = 'market_marketable_restriction'
    
    for i in result:
        for j in range(len(i['name'])):
            temp = 'A' + str(index + 1)
            ws[temp] = index
            temp = 'B' + str(index + 1)
            ws[temp] = index // 25 + 1
            temp = 'C' + str(index + 1)
            ws[temp] = index % 25
            temp = 'D' + str(index + 1)
            ws[temp] = i['name'][j]
            temp = 'E' + str(index + 1)
            ws[temp] = i['market_name'][j]
            temp = 'F' + str(index + 1)
            ws[temp] = i['market_hash_name'][j]
            temp = 'G' + str(index + 1)
            ws[temp] = i['quality'][j]
            temp = 'H' + str(index + 1)
            ws[temp] = i['rarity'][j]
            temp = 'I' + str(index + 1)
            ws[temp] = i['hero'][j]
            temp = 'J' + str(index + 1)
            ws[temp] = i['types'][j]
            temp = 'K' + str(index + 1)
            ws[temp] = i['slot'][j]
            temp = 'L' + str(index + 1)
            ws[temp] = i['tradable'][j]
            temp = 'M' + str(index + 1)
            ws[temp] = i['marketable'][j]
            temp = 'N' + str(index + 1)
            ws[temp] = i['currency'][j]
            temp = 'O' + str(index + 1)
            ws[temp] = i['commodity'][j]
            temp = 'P' + str(index + 1)
            ws[temp] = i['market_tradable_restriction'][j]
            temp = 'Q' + str(index + 1)
            ws[temp] = i['market_marketable_restriction'][j]
            index += 1
    
    wb.save('城市编码汇总.xlsx')




