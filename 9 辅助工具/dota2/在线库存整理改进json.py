# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 16:44:45 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts

实现了利用本地cookie文件的登录
"""

import requests
import json
from openpyxl import Workbook

def get_data(assetid, isInit = True, step = 1000):
    headers = {
    'Cache-Control': 'no-cache',
    'Connection': 'Transfer-Encoding',
    'Connection': 'keep-alive',
    'Content-Security-Policy': 'default-src blob: data: https: \'unsafe-inline\' \'unsafe-eval\'; script-src \'self\' \'unsafe-inline\' \'unsafe-eval\' https://steamcommunity-a.akamaihd.net/ https://api.steampowered.com/ *.google-analytics.com https://www.google.com https://www.gstatic.com https://apis.google.com; object-src \'none\'; connect-src \'self\' https://api.steampowered.com/ https://store.steampowered.com/ wss://community.steam-api.com/websocket/ *.google-analytics.com http://127.0.0.1:27060 ws://127.0.0.1:27060; frame-src \'self\' steam: https://store.steampowered.com/ https://www.youtube.com https://www.google.com https://sketchfab.com https://player.vimeo.com;',
    'Content-Type': 'application/json; charset=utf-8',
    'Date': 'Sat, 26 May 2018 13:09:57 GMT',
    'Expires': 'Mon, 26 Jul 1997 05:00:00 GMT',
    'Server': 'Apache',
    'Strict-Transport-Security': 'max-age=3600',
    'Transfer-Encoding': 'chunked',
    'X-Frame-Options': 'SAMEORIGIN'
    }
    pre_url = 'https://steamcommunity.com/inventory/76561198129439885/570/2?l=schinese&count='
    if isInit:
        url = pre_url + str(step)
    else:
        url = pre_url + str(step) + '&start_assetid=' + assetid
    try:
        response = requests.get(url, headers = headers)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None
        
def check(json_data):
    assets = json_data['assets']
    descriptions = json_data['descriptions']
    temp2 = descriptions[0]
    temp = {}
    for key, value in temp2.items():
        if type(value) == int:
            temp[key] = -1
        elif type(value) == str:
            temp[key] = ''
        elif type(value) == list:
            temp[key] = []
    index2 = 0
    for index1 in range(len(assets)):
        classid = assets[index1]['classid']
        instanceid = assets[index1]['instanceid']
        if index2 >= len(descriptions):
            temp2 = temp
            temp2['appid'] = assets[index1]['appid']
            temp2['classid'] = assets[index1]['classid']
            temp2['instanceid'] = assets[index1]['instanceid']
            descriptions.insert(index1, temp2)
            index2 += 1
        else:
            classid2 = descriptions[index2]['classid']
            instanceid2 = descriptions[index2]['instanceid']
            if classid != classid2 or instanceid != instanceid2:
                #print(index1)
                temp2 = temp
                temp2['appid'] = assets[index1]['appid']
                temp2['classid'] = assets[index1]['classid']
                temp2['instanceid'] = assets[index1]['instanceid']
                descriptions.insert(index1, temp2)
            index2 += 1
    
    return descriptions
    

if __name__ == '__main__':
    
    step = 1000
    isInit = True
    assetid = 'start'
    num = 0
    result = []
    
    while True:
        html = get_data(assetid, isInit, step)
        if html == None:
            print('运行错误！')
            break
        else:
            isInit = False
            json_data = json.loads(html)
            descriptions = check(json_data)
            result.append(descriptions)
            str_out = str(num * step) + ' to ' +  str(num * step + step) + 'Done!'
            print(str_out)
            num += 1
            if 'last_assetid' in json_data:
                assetid = json_data['last_assetid']
            else:
                break                 
    
    
    index = 1
    wb = Workbook()
    ws = wb.active
    ws.title = '刀塔在线库存'
    ws['A1'] = '序号'
    ws['B1'] = '页数'
    ws['C1'] = '子序号'
    ws['D1'] = '名称'
    ws['E1'] = '市场中文名称'
    ws['F1'] = '市场英文名称'
    ws['G1'] = '类型'
    ws['H1'] = 'classid'
    ws['I1'] = 'instanceid'
    ws['J1'] = '品质'
    ws['K1'] = '稀有度'
    ws['L1'] = '英雄'
    ws['M1'] = '属性'
    ws['N1'] = '槽位'
    
    ws['O1'] = '可交易'
    ws['P1'] = '可出售'
    ws['Q1'] = 'currency'
    ws['R1'] = 'commodity'
    ws['S1'] = 'market_tradable_restriction'
    ws['T1'] = 'market_marketable_restriction'

    
    for i in result:
        for j in i:
            temp = 'A' + str(index + 1)
            ws[temp] = index
            temp = 'B' + str(index + 1)
            ws[temp] = index // 25 + 1
            temp = 'C' + str(index + 1)
            temp_mod = index % 25
            if temp_mod == 0:
                temp_mod = 25
            ws[temp] = temp_mod
            temp = 'D' + str(index + 1)
            ws[temp] = j['name']
            temp = 'E' + str(index + 1)
            ws[temp] = j['market_name']
            temp = 'F' + str(index + 1)
            ws[temp] = j['market_hash_name']
            temp = 'G' + str(index + 1)
            ws[temp] = j['type']
            temp = 'H' + str(index + 1)
            ws[temp] = j['classid']
            temp = 'I' + str(index + 1)
            ws[temp] = j['instanceid']
            if len(j['tags']) == 0:
                temp = 'J' + str(index + 1)
                ws[temp] = 'NULL'
                temp = 'K' + str(index + 1)
                ws[temp] = 'NULL'
                temp = 'L' + str(index + 1)
                ws[temp] = 'NULL'
                temp = 'M' + str(index + 1)
                ws[temp] = 'NULL'
                temp = 'N' + str(index + 1)
                ws[temp] = 'NULL'
            else:
                temp = 'J' + str(index + 1)
                ws[temp] = j['tags'][0]['localized_tag_name']
                temp = 'K' + str(index + 1)
                ws[temp] = j['tags'][1]['localized_tag_name']
                temp = 'L' + str(index + 1)
                ws[temp] = j['tags'][4]['localized_tag_name']
                temp = 'M' + str(index + 1)
                ws[temp] = j['tags'][2]['localized_tag_name']
                temp = 'N' + str(index + 1)
                ws[temp] = j['tags'][3]['localized_tag_name']                
            temp = 'O' + str(index + 1)
            ws[temp] = j['tradable']
            temp = 'P' + str(index + 1)
            ws[temp] = j['marketable']
            temp = 'Q' + str(index + 1)
            ws[temp] = j['currency']
            temp = 'R' + str(index + 1)
            ws[temp] = j['commodity']
            temp = 'S' + str(index + 1)
            ws[temp] = j['market_tradable_restriction']
            temp = 'T' + str(index + 1)
            ws[temp] = j['market_marketable_restriction']
            
            index += 1
    
    wb.save('在线json库存管理.xlsx')




