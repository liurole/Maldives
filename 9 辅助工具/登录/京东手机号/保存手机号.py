# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 22:03:08 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts
"""
import requests
import random
import json
import openpyxl
import time
from urllib.parse import urlencode

# 利用request，导入cookies，header进行关键词网页搜索
def get_html(provinceId, cityId, rand_t):
    headers = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate, sdch, br',
    'Accept-Language': 'zh-CN,zh;q=0.8',
    'Connection': 'keep-alive',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest'
    }
    cookies = {
    'wlfstk_smdl': '3hoepa4lcounjmow84bzqe8myaar1z85',
    '__jda': '122270672.1519002416272376321166.1519002416.1519002416.1519013090.2',
    '__jdb': '122270672.3.1519002416272376321166|2.1519013090',
    '__jdc': '122270672',
    '__jdv': '122270672|direct|-|none|-|1519002416279',
    '__jdu': '1519002416272376321166',
    '_jrda': '2',
    '3AB9D23F7A4B3C9B': '6YB5BC53U3WV5UIV2ORHZALLHVJSVSRZTW4R5DL4MMK3VIQOA2NBJ65NGVXJAHCYYHAQRRZILNJHPKZ4PFC6TKMHZU',
    '_jrdb': '1519013090510',
    'TrackID': '1hXKTqHkuN9pu3MkF3ow4xFYM7QnWPUCAxNxrgvlsUluQW5yusKBdhi1Yl8kKMRPI7weRU_bQH1VnBV3cqmpRYtcKFC6ydjmelzGMkKnPjCCuqPknuwSYXNZPwxAVTgWQ',
    'pinId': '_e5GwNYT1_w',
    'pin': 'liurole',
    'unick': 'liurole',
    'thor': '8D05DD8BF36C31575232BE5C7A68E73479C020E0866666BDE90E3AF31413A1D5779B3270C8163387E700533463D662BF667EACB61CE1ED7DE3ADE85219B08950F4AFA7CFFBB46B6CA8D6CDB5C62C5E687BA0EB0742D9771742DB43035C0FAEA8B0B51A872BEC4E4D594313476482D2C6E906EE9FC4E6180D63B767F0BA077793A02526475436678D94565BA611B013B9',
    '_tp': 'SrQxwcWuTKQNSdZx9UnmXQ%3D%3D',
    '_pst': 'liurole',
    'ceshi3.com': '000'
    }
    data = {
    'cardWid': '5233026',
    'serviceOperatorId': '4',
    'bType': '86',
    'provinceId': str(provinceId),
    'cityId': str(cityId),
    'ownerInfoFlag': '0',
    'enc': 'C981E6CE8264F440',
    'numberSegment': '0',
    'numberRule': '0',
    'keyWord': '',
    't': str(rand_t),
    'pageIndex': '0'
    }
    params = urlencode(data)
    base = 'https://reselleve.jd.com/queryPhoneNumList.action'
    url = base + '?' + params
    try:
        response = requests.get(url, cookies = cookies, headers = headers)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

# 获取所有手机号
def get_all_nums(html):
    if html != '[]' and not(html is None):
        data = json.loads(html)
        hit = 0
        for i in data:
            if i['cardKey'] in city:
                hit += 1
                if hit > 14:
                    return True
            else:
                num.append(i['cardKey'])
                print('------' + i['cardKey'])
        return False
    else:
        return True  

if __name__ == '__main__':
    
    wb = openpyxl.reader.excel.load_workbook('城市编码汇总.xlsx')
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])

    # 获取表头
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数 
    city = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        city.append(app)
    
    print('Read Done!')

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '京东手机号'
    ws_out['A1'] = '省份'
    ws_out['B1'] = '城市'
    ws_out['C1'] = '手机号'
    index = 2
    num = []
    
    for i in range(300, len(city)):
        provinceId = city[i]['provinceId']
        cityId = city[i]['cityId']
        num.clear()
        bestop = False
        while not bestop:
            t = random.random() 
            html = get_html(provinceId, cityId, t)
            bestop = get_all_nums(html)
#            if len(num) < 100:
#                bestop = get_all_nums(html)
#            else:
#                bestop = True
        print(city[i]['省份'] + city[i]['城市'] + 'Done')
        time.sleep(5)
        if num != []:
            for j in range(len(num)):
                temp = 'A' + str(index)
                ws_out[temp] = city[i]['省份']
                temp = 'B' + str(index)
                ws_out[temp] = city[i]['城市']
                temp = 'C' + str(index)
                ws_out[temp] = num[j]
                index += 1
      
    wb_out.save('手机号汇总.xlsx')                
    
