# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 22:03:08 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts
"""
import requests
import json
from openpyxl import Workbook
from json.decoder import JSONDecodeError
from urllib.parse import urlencode

idp = []
idc = []
province = []
city = []
dicts = {
    '1': '北京',
    '2': '上海',
    '3': '天津',
    '4': '重庆',
    '5': '河北',
    '6': '山西',
    '7': '河南',
    '8': '辽宁',
    '9': '吉林',
    '10': '黑龙江',
    '11': '内蒙古',
    '12': '江苏',
    '13': '山东',
    '14': '安徽',
    '15': '浙江',
    '16': '福建',
    '17': '湖北',
    '18': '湖南',
    '19': '广东',
    '20': '广西',
    '21': '江西',
    '22': '四川',
    '23': '海南',
    '24': '贵州',
    '25': '云南',
    '26': '西藏',
    '27': '陕西',
    '28': '甘肃',
    '29': '青海',
    '30': '宁夏',
    '31': '新疆' }

# 利用request，导入cookies，header进行省份搜索
def get_html(provinceId):
    headers = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate, sdch, br',
    'Accept-Language': 'zh-CN,zh;q=0.8',
    'Connection': 'keep-alive',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest'
    }
    cookies = {
    'wlfstk_smdl': '3hoepa4lcounjmow84bzqe8myaar1z85',
    '__jda': '122270672.1519002416272376321166.1519002416.1519002416.1519013090.2',
    '__jdb': '122270672.3.1519002416272376321166|2.1519013090',
    '__jdc': '122270672',
    '__jdv': '122270672|direct|-|none|-|1519002416279',
    '__jdu': '1519002416272376321166',
    '_jrda': '2',
    '3AB9D23F7A4B3C9B': '6YB5BC53U3WV5UIV2ORHZALLHVJSVSRZTW4R5DL4MMK3VIQOA2NBJ65NGVXJAHCYYHAQRRZILNJHPKZ4PFC6TKMHZU',
    '_jrdb': '1519013090510',
    'TrackID': '1hXKTqHkuN9pu3MkF3ow4xFYM7QnWPUCAxNxrgvlsUluQW5yusKBdhi1Yl8kKMRPI7weRU_bQH1VnBV3cqmpRYtcKFC6ydjmelzGMkKnPjCCuqPknuwSYXNZPwxAVTgWQ',
    'pinId': '_e5GwNYT1_w',
    'pin': 'liurole',
    'unick': 'liurole',
    'thor': '8D05DD8BF36C31575232BE5C7A68E73479C020E0866666BDE90E3AF31413A1D5779B3270C8163387E700533463D662BF667EACB61CE1ED7DE3ADE85219B08950F4AFA7CFFBB46B6CA8D6CDB5C62C5E687BA0EB0742D9771742DB43035C0FAEA8B0B51A872BEC4E4D594313476482D2C6E906EE9FC4E6180D63B767F0BA077793A02526475436678D94565BA611B013B9',
    '_tp': 'SrQxwcWuTKQNSdZx9UnmXQ%3D%3D',
    '_pst': 'liurole',
    'ceshi3.com': '000'
    }
    data = {
    'provinceId': str(provinceId),
    'serviceOperatorId': '4',
    'bType': '86',
    'cardWid': '5233026',
    }
    params = urlencode(data)
    base = 'https://reselleve.jd.com/changeProvince.action'
    url = base + '?' + params
    try:
        response = requests.get(url, cookies = cookies, headers = headers)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

# 获取城市编码
def get_all_nums(html, provinceId):
    try:
        data = json.loads(html)
        if data:
            for i in data.get('citys'):
                province.append(dicts[str(provinceId)])
                city.append(i['name'])
                idp.append(provinceId)  
                idc.append(i['id'])  
    except JSONDecodeError:
        pass    
    
if __name__ == '__main__':
    
    wb = Workbook()
    ws = wb.active
    ws.title = '城市编码汇总'
    ws['A1'] = 'provinceId'
    ws['B1'] = 'cityId'
    ws['C1'] = '省份'
    ws['D1'] = '城市'
    
    for i in range(len(dicts)):
        provinceId = i + 1
        html = get_html(provinceId)
        get_all_nums(html, provinceId)
   
    for i in range(len(idc)):
        temp = 'A' + str((i + 2))
        ws[temp] = idp[i]
        temp = 'B' + str((i + 2))
        ws[temp] = idc[i]
        temp = 'C' + str((i + 2))
        ws[temp] = province[i]
        temp = 'D' + str((i + 2))
        ws[temp] = city[i]
        
    wb.save('城市编码汇总.xlsx')
    