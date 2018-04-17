# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 22:03:08 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts
"""
import requests
import json
import openpyxl
import time
from urllib.parse import urlencode

# 利用request，导入cookies，header进行关键词网页搜索
def get_html(cityId, cityName, index = 1):
    headers = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Cookie': 'woniu_test=woniu_test; closeTime=1523357409336; PHPSESSID=qt29n6cet075sv6v11bdsljln3; loginaccount=LIUROLE; sso.login.account=LIUROLE; nickname=LIUROLE; sex=0; Hm_lvt_cf6206191f9fb231bed0677b2bb51632=1523357399; WNAD_wd=-; WNAD_sd=-; WNAD_td=-; WNAD_ud=9F638478D796CC5AB96C3DBE029A8712; account=LIUROLE; naid=64034283; Hm_lpvt_cf6206191f9fb231bed0677b2bb51632=1523407395; 1TJ_P_28_=1523407395033_1523407395033_mall.snail.com; 1TJ_P_28__MEDIA=3__1523407395033_; 1TJ_D_=1523407395033_1523407395033_mall.snail.com; 1TJ_D__MEDIA=3__1523407395033_; mall_token=ea28d284ed8fdce9f4e64352bb6d0d7c',
    'Host': 'mall.snail.com',
    'Referer': 'http://mall.snail.com/purchase/newPick?goods_id=76&amount=1&mix_id=0&treaty_id=0&package_id=93&fcode=&f=&token=b04100ceb8aa0fdccdae600dede4dad2',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.108 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest'
    }
    cookies = {
    '1TJ_D_': '1523407395033_1523407395033_mall.snail.com',
    '1TJ_D__MEDIA': '3__1523407395033_',
    '1TJ_P_28_': '1523407395033_1523407395033_mall.snail.com',
    '1TJ_P_28__MEDIA': '3__1523407395033_',
    'Hm_lpvt_cf6206191f9fb231bed0677b2bb51632': '1523407395',
    'Hm_lvt_cf6206191f9fb231bed0677b2bb51632': '1523357399',
    'PHPSESSID': 'qt29n6cet075sv6v11bdsljln3',
    'WNAD_sd': '-',
    'WNAD_td': '-',
    'WNAD_ud': '9F638478D796CC5AB96C3DBE029A8712',
    'WNAD_wd': '-',
    'account': 'LIUROLE',
    'closeTime': '1523357409336',
    'loginaccount': 'LIUROLE',
    'mall_token': 'ea28d284ed8fdce9f4e64352bb6d0d7c',
    'naid': '64034283',
    'nickname': 'LIUROLE',
    'sex': '0',
    'sso.login.account': 'LIUROLE',
    'woniu_test': 'woniu_test'
    }
    data = {
    'package_id': '93',
    'type_id': '0',
    'city_code': cityId,
    'city_name': cityName,
    'pro_name': '江苏',
    'pretty_type': '',
    'page': str(index),
    'number_custom': '',
    'totalPage': '1',
    'mix_id': '0',
    'treaty_id': '0'
    }
    params = urlencode(data)
    base = 'http://mall.snail.com/purchase/Newpickinquiry'
    url = base + '?' + params
    try:
        response = requests.get(url, cookies = cookies, headers = headers)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

# 获取所有手机号
def get_all_nums(html):
    if html != '[]' and not(html is None):
        data = json.loads(html)
        page_total = data['page_total']
        for i in range(1, page_total): 
            html_temp = get_html(cityId, cityName, i)
            data_temp = json.loads(html_temp)
            for j in data_temp['number_data']:
                num.append(j['serial_number'])

        return False
    else:
        return True  

if __name__ == '__main__':
    
    wb = openpyxl.reader.excel.load_workbook('城市编码汇总.xlsx')
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])

    # 获取表头
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数 
    city = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        city.append(app)
    
    print('Read Done!')

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '蜗牛手机号'
    ws_out['A1'] = '省份'
    ws_out['B1'] = '城市'
    ws_out['C1'] = '手机号'
    index = 2
    num = []
    
    for i in range(len(city)):   # len(city)
        cityId = city[i]['cityId']
        cityName = city[i]['城市']
        num.clear()
        html = get_html(cityId, cityName)
        get_all_nums(html)

        print(city[i]['省份'] + city[i]['城市'] + 'Done')
        time.sleep(5)
        if num != []:
            for j in range(len(num)):
                temp = 'A' + str(index)
                ws_out[temp] = city[i]['省份']
                temp = 'B' + str(index)
                ws_out[temp] = city[i]['城市']
                temp = 'C' + str(index)
                ws_out[temp] = num[j]
                index += 1
      
    wb_out.save('手机号汇总.xlsx')                
    
