# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 22:03:08 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts
"""

import openpyxl

if __name__ == '__main__':
    
    wb = openpyxl.reader.excel.load_workbook('手机号汇总6.xlsx')
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])

    # 获取表头
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数 
    nums = []  
    tels = []
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        nums.append(app)
        tels.append(app['手机号'])
        if rownum % 100 == 0:
            print('-----------' + str(rownum))
    
    print('Read Done!')

    num_d = {}
    num_s = {}
    num_o = {}
    num_abc = {}
    num_b = {}
    digits = {}
    for i in tels:
        i = str(i)
        kind_s = 0
        kind_o = 0
        kind_abc = 0
        kind_b = 0
        digits.clear()
        for j in range(10):
            temp = i.count(str(j))
            if temp != 0:
                digits[str(j)] = temp
            if j == 0:
                temp = '000'
            else:                
                temp = str(j * 111)
            if temp in i:
                kind_s = 1
        for j in range(8): 
            if j == 0:
                temp = '012'
            else:
                temp = str(j * 100 + (j + 1) * 10 + (j + 2))
            if temp in i:
                kind_o = 1
        for j in range(6): 
            temp = i[j:(j+3)]
            k = i.find(temp, (j+3))
            if k != -1:
                kind_abc = 1
        if '0611' in i:
            kind_b = 1
        
        num_d[i] = len(digits)
        num_s[i] = kind_s
        num_o[i] = kind_o
        num_abc[i] = kind_abc
        num_b[i] = kind_b

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '蜗牛手机号'
    ws_out['A1'] = '手机号'
    ws_out['B1'] = '位数'
    ws_out['C1'] = '豹子'
    ws_out['D1'] = '连续'
    ws_out['E1'] = '重复'
    ws_out['F1'] = '生日'
    index = 2
    
    for i in range(len(tels)):
        temp = 'A' + str(i + 2)
        ws_out[temp] = tels[i]
        temp = 'B' + str(i + 2)
        ws_out[temp] = num_d[str(tels[i])]
        temp = 'C' + str(i + 2)
        ws_out[temp] = num_s[str(tels[i])]    
        temp = 'D' + str(i + 2)
        ws_out[temp] = num_o[str(tels[i])]  
        temp = 'E' + str(i + 2)
        ws_out[temp] = num_abc[str(tels[i])]          
        temp = 'F' + str(i + 2)
        ws_out[temp] = num_b[str(tels[i])]        
        
    wb_out.save('out6.xlsx')                
    
