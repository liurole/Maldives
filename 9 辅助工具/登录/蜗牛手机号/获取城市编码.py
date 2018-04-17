# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 22:03:08 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts
"""
import requests
import json
from openpyxl import Workbook
from json.decoder import JSONDecodeError

idp = []
idc = []
province = []
city = []
dicts = {
    '11': '北京',
    '31': '上海',
    '13': '天津',
    '34': '江苏',
    '36': '浙江',
    '51': '广东',
    '17': '山东',
    '30': '安徽',
    '50': '海南',
    '18': '河北',
    '19': '山西',
    '38': '福建',
    '59': '广西',
    '71': '湖北',
    '74': '湖南',
    '75': '江西',
    '76': '河南',
    '81': '四川',
    '83': '重庆',
    '84': '陕西',
    '85': '贵州',
    '86': '云南',
    '87': '甘肃',
    '88': '宁夏',
    '90': '吉林',
    '91': '辽宁',
    '97': '黑龙江',
    '10': '内蒙古' }

# 利用request，导入cookies，header进行省份搜索
def get_html(provinceId):
    headers = {
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Content-Length': '16',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Cookie': 'woniu_test=woniu_test; closeTime=1523357409336; PHPSESSID=qt29n6cet075sv6v11bdsljln3; account=LIUROLE; loginaccount=LIUROLE; naid=64034283; sso.login.account=LIUROLE; nickname=LIUROLE; sex=0; Hm_lvt_cf6206191f9fb231bed0677b2bb51632=1523357399; WNAD_wd=-; WNAD_sd=-; WNAD_td=-; Hm_lpvt_cf6206191f9fb231bed0677b2bb51632=1523358041; WNAD_ud=9F638478D796CC5AB96C3DBE029A8712; 1TJ_P_28_=1523357398311_1523358040625_mall.snail.com; 1TJ_P_28__MEDIA=3__1523358040625_; 1TJ_D_=1523357398311_1523358040625_mall.snail.com; 1TJ_D__MEDIA=3__1523358040625_; mall_token=7eabc09ace19ff35751258ce9ddfb28c',
    'Host': 'mall.snail.com',
    'Origin': 'http://mall.snail.com',
    'Referer': 'http://mall.snail.com/purchase/newPick?goods_id=76&amount=1&mix_id=0&treaty_id=0&package_id=93&fcode=&f=&token=e741070ff7040ba94d446ee5f303f0b3',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.108 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest'
    }
    cookies = {
    '1TJ_D_': '1523357398311_1523357398311_mall.snail.com',
    '1TJ_D__MEDIA': '3__1523357398311_',
    '1TJ_P_28_': '1523357398311_1523357398311_mall.snail.com',
    '1TJ_P_28__MEDIA': '3__1523357398311_',
    'Hm_lpvt_cf6206191f9fb231bed0677b2bb51632': '1523357399',
    'Hm_lvt_cf6206191f9fb231bed0677b2bb51632': '1523357399',
    'PHPSESSID': 'qt29n6cet075sv6v11bdsljln3',
    'WNAD_sd': '-',
    'WNAD_td': '-',
    'WNAD_ud': '-',
    'WNAD_wd': '-',
    'account': 'LIUROLE',
    'closeTime': '1.52336E+12',
    'loginaccount': 'LIUROLE',
    'mall_token': 'd1c4acdd3677660c3782a034809f7085',
    'naid': '64034283',
    'nickname': 'LIUROLE',
    'sex': '0',
    'sso.login.account': 'LIUROLE',
    'woniu_test': 'woniu_test'
    }
    datas = {
    'province_code': provinceId
    }
    url = 'http://mall.snail.com/purchase/getCity'
    try:
        response = requests.post(url, headers = headers, cookies = cookies, data = datas)
        if response.status_code == 200:
            return response.text
        return None
    except ConnectionError:
        print('Error occurred')
        return None

# 获取城市编码
def get_all_nums(html, provinceId):
    try:
        data = json.loads(html)
        if data:
            for i in data['data']:
                province.append(dicts[provinceId])
                city.append(i['name'])
                idp.append(i['province_code'])  
                idc.append(i['code'])  
    except JSONDecodeError:
        pass    
    
if __name__ == '__main__':
    
    wb = Workbook()
    ws = wb.active
    ws.title = '城市编码汇总'
    ws['A1'] = 'provinceId'
    ws['B1'] = 'cityId'
    ws['C1'] = '省份'
    ws['D1'] = '城市'
    
    for key,value in dicts.items():
        html = get_html(key)
        get_all_nums(html, key)
   
    for i in range(len(idc)):
        temp = 'A' + str((i + 2))
        ws[temp] = idp[i]
        temp = 'B' + str((i + 2))
        ws[temp] = idc[i]
        temp = 'C' + str((i + 2))
        ws[temp] = province[i]
        temp = 'D' + str((i + 2))
        ws[temp] = city[i]
        
    wb.save('城市编码汇总.xlsx')
    