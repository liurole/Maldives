# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 22:03:08 2018

@author: Se

下载chromedrive驱动 
使用Selenium需要选择一个调用的浏览器并下载好对应的驱动，本文使用chrome浏览器，当然也可以用FireFox等
http://www.seleniumhq.org/download/ 找到Google Chrome Driver链接
对应驱动放在python目录下面的scripts目录中，例如C:\ProgramData\Anaconda3\envs\python35\Scripts
"""
import requests
import re
import math
from openpyxl import Workbook

# 获取参数
def get_html(i_page):
    headers = {
    'Accept': 'text/html, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Content-Length': '335',
    'Content-Type': 'application/x-www-form-urlencoded',
    'Cookie': 'UM_distinctid=1628696a82b273-089b34a3a2f3cf-572f7b6e-144000-1628696a82cb14; CNZZDATA3874533=cnzz_eid%3D582286508-1522674683-%26ntime%3D1522674683; _qdda=3-1.1; _qddab=3-th27a5.jfia8pmg; _qddamta_2852151297=3-0; Hm_lvt_a936023e65b463d924c7377b8deb3d56=1522676444; Hm_lpvt_a936023e65b463d924c7377b8deb3d56=1522676444; __root_domain_v=.crd.cn; _qddaz=QD.xi9tm4.3rhqfh.jfia8pg3',
    'Host': 'www.crd.cn',
    'Origin': 'https://www.crd.cn',
    'Referer': 'https://www.crd.cn/diamond/wmin0.5wmax0.59/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.79 Safari/537.36 Edge/14.14393',
    'X-Requested-With': 'XMLHttpRequest'
    }
    cookies = {
    '__root_domain_v': '.crd.cn', 
    '_qdda': '3-1.1',
    '_qddab': '3-th27a5.jfia8pmg',
    '_qddamta_2852151297': '3-0',
    '_qddaz': 'QD.xi9tm4.3rhqfh.jfia8pg3 ',
    'CNZZDATA3874533': 'cnzz_eid%3D582286508-1522674683-%26ntime%3D1522674683 ',
    'Hm_lpvt_a936023e65b463d924c7377b8deb3d56': '1522676444',
    'Hm_lvt_a936023e65b463d924c7377b8deb3d56': '1522676444',
    'UM_distinctid': '1628696a82b273-089b34a3a2f3cf-572f7b6e-144000-1628696a82cb14'
    }
    data_s = {
    'wmin': '0.5',
    'wmax': '0.59',
    'pmin': '',
    'pmax': '',
    'color': 'E,F,G,H,I,J,K,E,F,G,H,E,F,G,H,',
    'cut': 'Ideal,EX,VG,GD,Fair,Ideal,EX,VG,Ideal,EX,VG,',
    'clarity': 'IF,VVS1,VVS2,VS1,VS2,SI1,SI2,IF,VVS1,VVS2,VS1,VS2,SI1,SI2,IF,VVS1,VVS2,VS1,VS2,SI1,SI2,',
    'polish': 'Ideal,EX,VG,GD,Fair,Ideal,EX,VG,Ideal,EX,VG,',
    'cert': 'GIA,IGI,HRD,GIA,IGI,HRD,GIA,IGI,HRD,',
    'symmetry': 'Ideal,EX,VG,GD,Fair,Ideal,EX,VG,Ideal,EX,VG,',
    'location': '1',
    'cpage': str(i_page),
    'order': '',
    'orderby': '',
    'keyword': '',
    'Shape': '圆形,圆形,圆形,',
    }
    url = 'https://www.crd.cn/Diamond/HDiamondSearch'
    response = requests.post(url, data = data_s)
    if response.status_code == 200:
        return response.text
    return None    

# 获取所有手机号
def get_all_nums():
    response = get_html(1)
    result = re.search('</div>\$\$\$(\d+)\$\$\$<a href', response, re.S)
    total_num = result.group(1)
    total_page = math.ceil(int(total_num) / 20)
    data_all = []
    index = 1
    
    for i in range(1, total_page + 1):
        print('page left\t' + str(total_page + 1 - i))
        response = get_html(i)
        results = re.findall('<span class=\'diamond3\'>(.*?)</sp.*?2\'>(.*?)</sp.*?2\'>(.*?)</sp.*?2\'>(.*?)</sp.*?2\'>(.*?)</sp.*?2\'>(.*?)</sp.*?2\'>(.*?)</sp.*?2\'>(.*?)</sp.*?none\'>(.*?)</sp.*?2\'>(.*?)</sp.*?2\'>(.*?)</span>', response, re.S)
        for result in results:
            temp = {}
            temp['编号'] = index
            temp['货号'] = result[0]
            temp['证书'] = result[1]
            temp['重量'] = result[2]
            temp['净度'] = result[3]
            temp['颜色'] = result[4]
            temp['切工'] = result[5]
            temp['抛光'] = result[6]
            temp['对称'] = result[7]
            temp['荧光'] = result[8]
            temp['价格'] = result[9]
            temp['所有地'] = result[10]
            index += 1
            data_all.append(temp)
    return data_all  

if __name__ == '__main__':
    
    data = get_all_nums()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '钻石参数汇总'
    ws['A1'] = '编号'
    ws['B1'] = '货号'
    ws['C1'] = '证书'
    ws['D1'] = '重量'
    ws['E1'] = '净度'
    ws['F1'] = '颜色'
    ws['G1'] = '切工'
    ws['H1'] = '抛光'
    ws['I1'] = '对称'
    ws['J1'] = '荧光'
    ws['K1'] = '价格'
    ws['L1'] = '所有地'

    for i in range(len(data)):
        temp = 'A' + str((i + 2))
        ws[temp] = data[i]['编号']
        temp = 'B' + str((i + 2))
        ws[temp] = data[i]['货号']
        temp = 'C' + str((i + 2))
        ws[temp] = data[i]['证书']
        temp = 'D' + str((i + 2))
        ws[temp] = data[i]['重量']
        temp = 'E' + str((i + 2))
        ws[temp] = data[i]['净度']
        temp = 'F' + str((i + 2))
        ws[temp] = data[i]['颜色']
        temp = 'G' + str((i + 2))
        ws[temp] = data[i]['切工']
        temp = 'H' + str((i + 2))
        ws[temp] = data[i]['抛光']
        temp = 'I' + str((i + 2))
        ws[temp] = data[i]['对称']
        temp = 'J' + str((i + 2))
        ws[temp] = data[i]['荧光']
        temp = 'K' + str((i + 2))
        ws[temp] = data[i]['价格']
        temp = 'L' + str((i + 2))
        ws[temp] = data[i]['所有地']

    wb.save('钻石参数汇总.xlsx')
